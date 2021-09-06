import openpyxl as xl
from openpyxl.chart import BarChart, Reference



def process_workbook(filename):
    # For loading the workbook
    wb = xl.load_workbook(filename)
    # For loading the sheet
    sheet = wb["Sheet1"]
    # For iterating over the sheet cell wise
    # 1st Method
    # cell = sheet["a1"]
    # print(cell.value)
    # 2nd Method
    cell = sheet.cell(1,1)
    # print(cell.value)

    # To know the max number of rows in sheet
    # print(sheet.max_row)

    # Updating the values and storing in a new column
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    # Saving the file with a different name
    wb.save(filename)

file_name = input("Enter file name : ")
process_workbook(file_name)